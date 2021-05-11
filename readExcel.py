from openpyxl import load_workbook
from openpyxl import Workbook
import app

def read_excel(i,j,excel_name):
    #excel_name은 엑셀파일 이름을 가져온 것이다.
    # data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = load_workbook("static/excel/"+ excel_name, data_only=True)
    # 시트 이름으로 불러오기
    load_ws = load_wb['Sheet']

    cell_obj = load_ws.cell(row=i, column=j)
    print(cell_obj.value)
    if cell_obj.value != None:
        return cell_obj.value
    else:
        return 0


